"""Formatted XLSX review workbook exporter with grouped color-coded blocks."""

from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from exporters.csv_exporter import CsvExporter


class XlsxReviewExporter:
    BLOCK_COLORS = {
        "Core": "7030A0",          # purple
        "Phone": "31869B",         # teal
        "Employment": "1F4E78",    # dark blue
        "Education": "C69214",     # gold
        "EducationAlt": "D6A800",  # alternate gold
        "Military": "548235",      # green
        "Relationships": "C00000", # red
        "Protected": "EADCF8",     # light purple protected marker
    }

    def __init__(self, schema_analysis, remove_empty_columns=True):
        self.schema_analysis = schema_analysis
        self.remove_empty_columns = remove_empty_columns

    def export(self, constituents, output_file):
        csv_exporter = CsvExporter(self.schema_analysis, remove_empty_columns=self.remove_empty_columns)
        headers = csv_exporter.schema_builder.build_headers(self.schema_analysis)
        block_names = self._build_block_names()
        rows = [csv_exporter._build_row(c) for c in csv_exporter._sorted_constituents(constituents)]

        if self.remove_empty_columns:
            headers, rows, block_names = self._remove_empty_columns_with_blocks(headers, rows, block_names)

        wb = Workbook()
        ws = wb.active
        ws.title = "CT Process Review"
        ws.sheet_view.showGridLines = False
        ws.append(block_names)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        self._merge_block_headers(ws, block_names)
        self._format_sheet(ws, headers, block_names)
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

    def _degree_width(self):
        return 8 + (1 + self.schema_analysis.get("max_major_records", 0)) + (1 + self.schema_analysis.get("max_minor_records", 0)) + (1 + self.schema_analysis.get("max_attribute_records", 0) * 2)

    def _build_block_names(self):
        blocks = []
        blocks.extend(["Core"] * len(CsvExporter(self.schema_analysis).schema_builder.CORE_FIELDS))
        for _ in range(self.schema_analysis.get("max_phone_records", 0)):
            blocks.extend(["Phone"] * 3)
        for _ in range(self.schema_analysis.get("max_employment_records", 0)):
            blocks.extend(["Employment"] * 10)
        degree_width = self._degree_width()
        for degree_index in range(self.schema_analysis.get("max_degree_records", 0)):
            blocks.extend([f"Education {degree_index + 1}"] * degree_width)
        for _ in range(self.schema_analysis.get("max_military_records", 0)):
            blocks.extend(["Military"] * 7)
        for _ in range(self.schema_analysis.get("max_relationship_records", 0)):
            blocks.extend(["Relationships"] * 3)
        return blocks

    def _block_color(self, block_name):
        if str(block_name).startswith("Education"):
            try:
                number = int(str(block_name).split()[1])
            except Exception:
                number = 1
            return self.BLOCK_COLORS["Education"] if number % 2 == 1 else self.BLOCK_COLORS["EducationAlt"]
        return self.BLOCK_COLORS.get(block_name, "1F4E78")

    def _merge_block_headers(self, ws, block_names):
        if not block_names:
            return
        start = 1
        current = block_names[0]
        for idx in range(2, len(block_names) + 2):
            next_value = block_names[idx - 1] if idx <= len(block_names) else None
            if next_value != current:
                end = idx - 1
                if end > start:
                    ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
                ws.cell(row=1, column=start).value = current
                start = idx
                current = next_value

    def _format_sheet(self, ws, headers, block_names):
        group_font = Font(name="Aptos Narrow", size=10, color="FFFFFF", bold=True)
        header_font = Font(name="Aptos Narrow", size=10, color="FFFFFF", bold=True)
        body_font = Font(name="Aptos Narrow", size=9)
        thin_gray = Side(style="thin", color="D9E2F3")
        medium_white = Side(style="medium", color="FFFFFF")
        warning_fill = PatternFill("solid", fgColor="FCE4D6")
        warning_font = Font(name="Aptos Narrow", size=9, color="9C0006", bold=True)
        new_fill = PatternFill("solid", fgColor="E2F0D9")
        existing_fill = PatternFill("solid", fgColor="DDEBF7")
        protected_fill = PatternFill("solid", fgColor=self.BLOCK_COLORS["Protected"])
        ws.freeze_panes = "C3"
        ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"

        for col_idx, block_name in enumerate(block_names, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = PatternFill("solid", fgColor=self._block_color(block_name))
            cell.font = group_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(right=medium_white)

        for col_idx, header in enumerate(headers, start=1):
            block_name = block_names[col_idx - 1]
            cell = ws.cell(row=2, column=col_idx)
            cell.fill = PatternFill("solid", fgColor=self._block_color(block_name))
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True)
            cell.border = Border(bottom=Side(style="medium", color="5B9BD5"), right=thin_gray)
            if header == "ESRPrimAlum":
                cell.fill = protected_fill
                cell.font = Font(name="Aptos", size=10, color="000000", bold=True)

        for row in ws.iter_rows(min_row=3):
            status = row[0].value
            if status == "NEW":
                row[0].fill = new_fill
            elif status == "EXISTING":
                row[0].fill = existing_fill
            for cell in row:
                cell.font = body_font
                cell.border = Border(bottom=thin_gray)
                cell.alignment = Alignment(vertical="top", wrap_text=False)
                cell.number_format = "@"
                if isinstance(cell.value, str) and "WARNING" in cell.value.upper():
                    cell.fill = warning_fill
                    cell.font = warning_font

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 42
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(header))
            sample_limit = min(ws.max_row, 80)
            for row_idx in range(3, sample_limit + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    max_len = max(max_len, min(len(str(value)), 42))
            ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 36))

    def _remove_empty_columns_with_blocks(self, headers, rows, block_names):
        if not rows:
            return headers, rows, block_names
        from config import PROTECTED_COLUMNS
        keep_indexes = []
        for idx, header in enumerate(headers):
            if header in PROTECTED_COLUMNS:
                keep_indexes.append(idx)
                continue
            if any(row[idx] not in (None, "") for row in rows):
                keep_indexes.append(idx)
        return [headers[i] for i in keep_indexes], [[row[i] for i in keep_indexes] for row in rows], [block_names[i] for i in keep_indexes]
